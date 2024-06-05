import os
import shutil
import openpyxl
import datetime 
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
from Model.fuma import Fuma


class ExcelLocal:
    def __init__(self):
        self.fuma = Fuma()
        self.file_mime_type = 'application/vnd.google-apps.spreadsheet'  # Tipo MIME de la Hoja de cálculo de Google
        self.ruta_origen = "C:/Users/Mario/Downloads/CONVERTIDOR EXCEL"
        self.ruta_destino = "C:/Users/Mario/Downloads/CONVERTIDOR EXCEL/Convertidos"
        self.contador = 0

    def mover_archivo(self, archivo):
        # Normalizar la ruta del archivo
        archivo = os.path.normpath(archivo)
        
        # Asegurarse de que la carpeta de destino existe
        if not os.path.exists(self.ruta_destino):
            os.makedirs(self.ruta_destino)

        # Construir la ruta completa del archivo de origen
        ruta_archivo_origen = os.path.join(self.ruta_origen, archivo)

        # Verificar que el archivo existe y que es un archivo
        if os.path.isfile(ruta_archivo_origen):
            # Construir la ruta completa del archivo de destino
            ruta_archivo_destino = os.path.join(self.ruta_destino, os.path.basename(archivo))
            # Mover el archivo
            shutil.move(ruta_archivo_origen, ruta_archivo_destino)
            print(f"Archivo {archivo} movido correctamente.")
        else:
            print(f"El archivo {archivo} no existe o no es un archivo válido.")

    def obtener_archivos(self):
        archivos = []
        # Obtener la lista de elementos en el directorio
        elementos = os.listdir(self.ruta_origen)
        # Filtrar solo los archivos
        for elemento in elementos:
            ruta_elemento = os.path.join(self.ruta_origen, elemento)
            if os.path.isfile(ruta_elemento):
                archivos.append(ruta_elemento)
        return archivos
    
    def obtener_id_desde_excel(self, archivo_excel):
        try:
            # Cargar el archivo de Excel
            wb = openpyxl.load_workbook(archivo_excel)
            # Seleccionar la primera hoja
            sheet = wb.active
            # Obtener el valor de la celda A2
            id_value = sheet[f'A{2+self.contador}'].value
            # Verificar si el valor es numérico
            if isinstance(id_value, (int, float)):
                # Establecer el ID en la instancia de Fuma
                self.fuma.set_id(int(id_value))
            else:
                print("El valor en la celda A2 no es numérico.")
        except Exception as e:
            print(f"Error al leer el archivo Excel: {e}")
    
    def obtener_fecha_desde_excel(self, archivo_excel):
        try:
            # Cargar el archivo de Excel
            wb = openpyxl.load_workbook(archivo_excel)
            # Seleccionar la primera hoja
            sheet = wb.active
            # Obtener el valor de la celda B2
            fecha_value = sheet[f'B{2+self.contador}'].value
            # Verificar si el valor no es None y es de tipo datetime
            if fecha_value is not None and isinstance(fecha_value, datetime.datetime):
                # Convertir el objeto datetime en una cadena de texto
                fecha_str = fecha_value.strftime('%d/%m/%Y')
                # Establecer la fecha en la instancia de Fuma
                self.fuma.set_fecha(fecha_str)
            else:
                print("La celda B2 está vacía o no contiene una fecha válida.")
        except Exception as e:
            print(f"Error al leer el archivo Excel: {e}")

    def obtener_solicitante_desde_excel(self, archivo_excel):
        strn = f'C{2+self.contador}'
        asignacion_value = self.get_string(archivo_excel, strn)
        if asignacion_value:
            self.fuma.set_solicitante(asignacion_value)

    def obtener_departamento_desde_excel(self, archivo_excel):
        strn = f'D{2+self.contador}'
        asignacion_value = self.get_string(archivo_excel, strn)
        if asignacion_value:
            self.fuma.set_departamento(asignacion_value)

    def obtener_asignacion_desde_excel(self, archivo_excel):
        strn = f'E{2+self.contador}'
        asignacion_value = self.get_string(archivo_excel, strn)
        if asignacion_value:
            self.fuma.set_asignacion(asignacion_value)
    
    def obtener_unidadnegocio_desde_excel(self, archivo_excel):
        strn = f'F{2+self.contador}'
        asignacion_value = self.get_string(archivo_excel, strn)
        if asignacion_value:
            self.fuma.set_unidadnegocio(asignacion_value)
    
    def obtener_proyecto_desde_excel(self, archivo_excel):
        strn = f'G{2+self.contador}'
        asignacion_value = self.get_string(archivo_excel, strn)
        if asignacion_value:
            self.fuma.set_proyecto(asignacion_value)
    
    def obtener_productos_desde_excel(self, archivo_excel):
        strn = f'H{2+self.contador}'
        asignacion_value = self.get_string(archivo_excel, strn)
        if asignacion_value:
            self.fuma.set_productos(asignacion_value)

    def get_string(self, archivo_excel, strn):
        try:
            # Cargar el archivo de Excel
            wb = openpyxl.load_workbook(archivo_excel)
            # Seleccionar la primera hoja
            sheet = wb.active
            # Obtener el valor de la celda especificada
            value = sheet[strn].value
            # Verificar si el valor es una cadena
            if isinstance(value, str):
                return value
            else:
                print("El valor en la celda", strn, "no es una cadena.")
        except Exception as e:
            print(f"Error al leer el archivo Excel: {e}")
        return None

    def revizar_celda_abajo(self, archivo_excel):
        try:
            # Cargar el archivo de Excel
            wb = openpyxl.load_workbook(archivo_excel)
            # Seleccionar la primera hoja
            sheet = wb.active
            
            # Incrementar el contador
            self.contador = self.contador + 1
            
            # Obtener el valor de la celda en la posición A2 + contador
            id_value = sheet[f'A{2+self.contador}'].value
            
            # Verificar si hay algún dato en la celda
            if id_value is not None:
                return False
            else:
                # Si no hay dato, reiniciar el contador y devolver False
                self.contador = 0
                return True
            
        except Exception as e:
            print(f"Error al leer el archivo Excel: {e}")
            return False



